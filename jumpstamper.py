#!/usr/bin/env python3

import sys
import ffmpeg
import openpyxl
import argparse
import copy
from dataclasses import dataclass
import pprint
import math





@dataclass 
class StamperArgs:
    '''
    Collects arguments from CLI, does some sanity checking.  Class instance should be a set of 
    values that are sane for passing into FFMPEG filters -- this includes things like durations can't 
    be negative, frame numbers can't be negative, you can't back a leadin to before the first frame, etc.
    '''

    input_file : str = ''
    output_file : str = ''
    stamp : bool = False
    slate_time : float = 0.0
    freeze_time : float = 0.0
    leadin_time : float = 0.0
    working_time : float = 0.0
    jump_time : float = 60.0    
    slate_frame : int = 0
    exit_frame : int = 0
    overlay_prof : str = 'default'
    encoder_prof : str = 'default'
    annotation : str = ''
    excel_sheet : str = None

    def numFrames(self, seconds: float) -> int: 
        return round(seconds * self.framerate)

    def numSecs(self, frames: int) -> float: 
        return (frames / self.framerate)


    def __post_init__(self): 
        '''a jump has five meaningful frame indices to keep:
            - slate, lead-in, exit, freeze, end
           and some useful durations, which we compute in both frames and time:
            - total jump,slate, lead-in, freeze
        ''' 
        self.vid_metadata = getVideoMetadata(self.input_file)
        self.framerate = getFrameRate(self.vid_metadata)
        # dict of durations in seconds (mostly from CLI)
        self.secs = {
            'slate' : self.slate_time,
            'leadin' : self.leadin_time,
            'working' : self.working_time,
            'freeze' : self.freeze_time,
            'jump' : self.jump_time,
            'total' : self.jump_time + self.freeze_time + self.leadin_time + self.slate_time
        }

        # the frame numbers for points of interest
        self.framenum = {
            'slate' : self.slate_frame,
            'leadin' : self.exit_frame - self.numFrames(self.leadin_time),
            'exit' : self.exit_frame,
            'freeze' : self.exit_frame + self.numFrames(self.working_time),
            'end' : self.exit_frame + self.numFrames(self.jump_time),
        }
        # compute the duration in FRAMES from the values we have in SECONDS
        self.frames = { k:self.numFrames(v) for k,v in self.secs.items() }
        self.sanity()

    def sanity(self):
        ''' run some sanity checks on the computed values '''

        # if the reqested leadin would be before the beginning of the jump,
        # reset the leadin frame to the start and adjust the duration
        if (self.framenum['leadin'] < 0):
            self.framenum['leadin'] = 0
            # the leadin duration is now the exit frame number comverted back to seconds
            self.secs['leadin'] = self.numSecs(self.framenum['exit'])



@dataclass
class StamperProfiles:
    '''
    the StamperProfile class contains the various settings (aka a "profile") for the various
    arguments and settings that we use in the main script.  The basic idea here is that the
    default settings are as portable as we can make them, and if people want to alter the various
    filter settings, they can do so here.  
    '''

    args : StamperArgs                  # the input StamperArgs
                                        # select which overlay profile: required

    def roundToMultOf(x, base=10):
        return base * round(x/base)

    def __post_init__(self):

        vid_metadata = getVideoMetadata(self.args.input_file)
        framerate = getFrameRate(vid_metadata)
        vid_h = vid_metadata['height']
        vid_w = vid_metadata['width']

        ''' the shared arguments for the drawtext filters.  Change if you need.  Font selection
        is particularly messy with FFMPEG, you can either provide an explicit font file (careful with
        OS-specific stuff here, or if your FFMPEG has fontconfig compiled in, just give it a 
        family and have the ffmpeg app determine what you use.  "font" is likely more portable
        if the support is compiled in...'''
        self.common_dt : dict = {                 
            # 'fontfile' : '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
            'font' : 'Arial',
            'rate' : framerate,
            'fontcolor' : 'yellow',
            'fontsize' : round(vid_metadata['height'] / 12 / 8) * 8,    # round to a multiple of 8
            'box' : 1,
            'boxcolor' : 'black@0.5',
            'boxborderw' : 2,
        }


        self.framectr_dt = {              # drawtext args for the frame counter
            **self.common_dt, 
            'y' : (vid_h * 0.9), 
            'x' : (vid_w * 0.2),
            'text' : 'in: %{frame_num} @ %{pts}',
            'start_number' : 0,
        }

        self.slate_trim = {               # trim filter for the slate
            'start_frame' : self.args.framenum['slate'],
            'end_frame' : self.args.framenum['slate'] + 1,
        }
        
        self.jump_trim = {                # trim filter for the main jump
            'start_frame': self.args.framenum['leadin'],   
            'end_frame': self.args.framenum['end'],
        }

        self.timer_dt = {                 # drawtext filter for timer
            **self.common_dt, 
            'y' : 0,
            'x' : 0,
        }

        annot_fontsize = round(vid_h / 16 / 8) * 8
        self.annot_dt = {                 # drawtext filter for annotation
            **self.common_dt,
            'fontsize': annot_fontsize,       # 1/16th of height rounded
            'y': vid_h - annot_fontsize,      # offset by the height from the bottom
            'x': 0,
            'text': self.args.annotation,
        }

        self.freeze_loop : dict = {             # loop filter for freezeframe
            # freeze starts here because we have to compensate for the 'trim'
            'start': self.args.framenum['freeze'] - self.args.framenum['leadin'] + 1,
            'loop': self.args.frames['freeze'],
            'size': 1,
        }
        
        fade_duration = 2
        self.jump_fade : dict = {               # fade filter for main jump
            'type' : 'out',
            'start_time' : self.args.secs['leadin'] + self.args.secs['jump'] + self.args.secs['freeze'] - fade_duration,
            'duration' : fade_duration,
        }
        
        self.fps : dict = {                     
            'fps' : getFrameRate(vid_metadata)                   # default to original value
        }
        
        self.scale : dict = {
            'width' : vid_w,
            'height' : vid_h,
        }

        self.input : dict = {
            'hide_banner': None,
            'benchmark': None,
            'an': None,
            'y': None,
        }

        self.output : dict = {                  # default output arguments
            'c:v': 'libx264',
            'b:v' : '10M',
            'format': 'mp4',
        }

        # outputs in 1080 @ 30fps
        if (self.args.encoder_prof == '1080_30'):
            self.scale = {'width': '-4', 'height': '1080'}
            self.fps = {'fps': 30}

        # quick - useful for iterative testing!
        if (self.args.encoder_prof == 'quick'):
            self.scale = {'width': '-4', 'height': '480'}
            # fps = {'fps': 25}
            self.output['crf']  = 30
            self.output['preset']  = 'veryfast'

        # good-quality output in h.265 but takes a while...
        if (self.args.encoder_prof == 'x265_high'):
            self.output['c:v'] = 'libx265'
            self.output['preset'] = 'slow'
            self.output['b:v'] = '10M'

        # HW accelerated h.264 (must have hardware/GPU support!)
        if (self.args.encoder_prof == 'qsv_h264'):
            self.output = {
            'c:v' : 'h264_qsv',
            'look_ahead' : '1',
            'look_ahead_depth' : '40',
            'b:v' : '5M',
            }
            self.input['hwaccel'] = 'qsv'
            self.input['loglevel'] = 'verbose'

        # HW accelerated HEVC/h.265 (must have hardware/GPU support!)
        if (self.args.encoder_prof == 'qsv_hevc'):
            self.output = {
                'c:v' : 'hevc_qsv',
                'preset' : '4',
                'b:v' : '10M',
            }
            self.input['hwaccel'] = 'qsv'
            self.input['loglevel'] = 'verbose'





        # the null profile does NOT inherit anything so we have to replicate here...
        # for debug/test only  ;-)
        if (self.args.encoder_prof == 'null'):
            self.output = {
                'format': 'null',
                'y': None,
            }





def jumpParser():
    parser = argparse.ArgumentParser()
    parser.add_argument("-i",   "--input_file", action="store", type=str, help="input video file")
    parser.add_argument("-o",   "--output_file", action="store", type=str,  help="output video file")
    parser.add_argument("-s",   "--stamp", action="store_true",  help="enable frame stamping")
    parser.add_argument("-st",  "--slate_time", action="store", type=float,   help="duration of the slate frame")
    parser.add_argument("-lt",  "--leadin_time", action="store", type=float,   help="start jump portion of video N seconds before exit")
    parser.add_argument("-wt",  "--working_time", action="store", type=float,   help="seconds of working (scoring) time")
    parser.add_argument("-jt",  "--jump_time", action="store", type=float,   help="length of the jump video")
    parser.add_argument("-ef",  "--exit_frame", action="store", type=int, help="frame number of the exit")
    parser.add_argument("-sf",  "--slate_frame", action="store", type=int,   help="frame number of the slate, 0 to disable")
    parser.add_argument("-ft",  "--freeze_time", action="store", type=float,   help="duration of the freeze, 0 to disable")
    parser.add_argument("-ovr", "--overlay_prof", action="store", type=str,   help="the overlay profile name")
    parser.add_argument("-enc", "--encoder_prof", action="store", type=str,   help="the encoder options profile name")
    parser.add_argument("-an",  "--annotation", action="store", type=str,   help="annotation string")
    parser.add_argument("-xlsx", "--excel_sheet", action="store", type=str, help="parse excel: header row required with arguments, each row is one jump")

    return parser


def getVideoMetadata(infile: str) -> dict:
    ''' returns a dict of metadata via ffprobe for the first VIDEO stream in input file. '''

    try:
        probe = ffmpeg.probe(infile) 
    except ffmpeg.Error as e:
        print(e.stderr, file=sys.stderr)
        sys.exit(1)

    video_stream = next((stream for stream in probe['streams'] if stream['codec_type'] == 'video'), None)
    if video_stream is None:
        print('No video stream found', file=sys.stderr)
        sys.exit(1)

    return video_stream


def getFrameRate(vid_metadata: dict) -> float:
    ''' returns a float with the probed frame rate from the video file. '''
    rate = vid_metadata['r_frame_rate'].split('/')
    # is it represented as a single integer value?
    if len(rate)==1:
        return float(rate[0])
    # or as a fraction (usually NTSC...)
    if len(rate)==2:
        return float(rate[0])/float(rate[1])
    return -1





def makeStamped(args: StamperArgs):
    
    settings = StamperProfiles(args)
    output = (
        ffmpeg
        .input(args.input_file, **settings.input)
        .filter('drawtext', **framecounter_args)
    )

    return output



def makeSlate(args: StamperArgs):
    
    settings = StamperProfiles(args)
    slate_trim_args = {
        'start_frame' : args.framenum['slate'],
        'end_frame' : args.framenum['slate'] + 1,
    }

    slate = (
        ffmpeg
        .input(args.input_file, **settings.input)
        .crop(width='0.8*in_w', height='ih', x='0.1*in_w', y=0)
        .trim(**slate_trim_args)    # spits out a single-frame stream
        .filter('loop', loop=args.frames['slate'], size=1, start=1)  # the first frame of the TRIMMED stream
        .setpts('N/FRAME_RATE/TB')      # fixup PTS for the looped frames
    )

    if args.secs['slate'] >= 2:
        fade_dur = 0.5
        fade_in_time = 0
        fade_out_time = args.secs['slate'] - fade_dur
        slate = (
            slate
            .filter('fade', type='in', start_time=fade_in_time, duration=fade_dur)
            .filter('fade', type='out', start_time=fade_out_time, duration=fade_dur)
        )

    return slate



# def makeMainJump(args: StamperArgs) -> ffmpeg.nodes.FilterableStream:
def makeMainJump(args: StamperArgs):

    settings = StamperProfiles(args)

    if (args.working_time == 0):          # set text to empty string, which causes drawtext to do nothing
        settings.timer_dt['text'] = ''         
    else:
        # this is an expression that drawtext can evaluate to get the timestamp in the format 
        # that we want.  See 'eif' 'trunc' 'abs' functions and 't' variable in docs.
        base_EIF_calc = '%{eif:(trunc(t-LEADIN)):d:2}.%{eif:abs((1M*(t-LEADIN)-1M*trunc(t-LEADIN))/10000):d:2}'
        settings.timer_dt['text'] = base_EIF_calc.replace('LEADIN', str(args.secs['leadin']))

    main_jump = (
        ffmpeg
        .input(args.input_file, **settings.input)
        .trim(**settings.jump_trim)
        .setpts('N/FRAME_RATE/TB')      # fixup PTS for the trimmed stream
        .crop(width='0.8*in_w', height='ih', x='0.1*in_w', y=0)
        .filter('drawtext', **settings.timer_dt)
        .filter('drawtext', **settings.annot_dt)
        .filter('loop', **settings.freeze_loop)  
        .setpts('N/FRAME_RATE/TB')      # fixup PTS for the looped frames
        .filter('fade', **settings.jump_fade)
    )
    return main_jump


# def joinAndPostFilters(args: StamperArgs, to_concat: list) -> ffmpeg.nodes.FilterableStream:
def joinAndPostFilters(args: StamperArgs, to_concat: list):

    settings = StamperProfiles(args)
    joined = (
        ffmpeg
        .concat(*to_concat, v=1, a=0)
        .filter('scale', **settings.scale)
        .filter('fps', **settings.fps)
        .node       # I still don't know what this does.. 
    )
    return joined




def processJumps(list_of_args: list):

    for cli_args in list_of_args:

        filter_args = StamperProfiles(cli_args)

        if cli_args.stamp:      # the stamp flag is set, so do that...
            framestamped = makeStamped(cli_args)
            to_output = joinAndPostFilters(cli_args, [framestamped])

        elif cli_args.slate_time == 0:      # there's no slate info from the CLI
            main_jump = makeMainJump(cli_args)
            to_output = joinAndPostFilters(cli_args, [main_jump])

        else:           # we have a slate to add up front...
            slate = makeSlate(cli_args)
            main_jump = makeMainJump(cli_args)
            to_output = joinAndPostFilters(cli_args, [slate, main_jump])

        if cli_args.encoder_prof == 'null':
            cli_args.output_file = '/dev/null'

        output = ffmpeg.output(to_output[0], cli_args.output_file, **filter_args.output)
        print (f"FFMPEG command string is: \n {' '.join(output.compile())}")
        output.run()


def jumpsFromXlsx(args: dict, known_args: list) -> list:
    ''' returns a list of dicts, where each dict is a set of CLI-equivalent arguments for a jump '''
    xls_file = args['excel_sheet']
    try:
        wb = openpyxl.load_workbook(xls_file)
    except Exception:
        print(f"failed to open file {xls_file}")
    ws = wb.active
    # TODO: assert that worksheet has at least two rows
    # TODO:assert that input_file and output_file are both in row 1


    jumpArgs = []
    for jumprow in ws.iter_rows(min_row=2):
        jump_args = {}
        for cell in jumprow:
            # param_key is the value from row 1 for that column
            param_key = ws.cell(row=1, column=cell.column).value
            param_value = cell.value
            # ignore/warn on any header values that aren't known ArgParse arguments
            if param_key in known_args:
                jump_args[param_key] = param_value
            else:
                print(f"ignoring an unknown header row value ({param_key}) in excel file: {xls_file}")
        if all (jump_args[k] is not None for k in ['input_file', 'output_file']):   # have to have in/out files, skip "empty" rows...
            jumpArgs.append(jump_args)

    return jumpArgs


def getCmdLineList(parser: argparse.ArgumentParser) -> list:
    '''
    if it's an excel sheet, return each set of arguments as an element in the list
    if it's a single jump from a set of CLI args, return a single element list containing the passed args
    '''
    known_parser_args = [action.dest for action in parser._actions]
    cleaned_jumps = []
    cli_args = vars(parser.parse_args())  
    if cli_args['excel_sheet'] is not None:
        jumpsToProcess = jumpsFromXlsx(cli_args, known_parser_args)
    else:
        jumpsToProcess = [cli_args]         # make the jump into a single item list
    

    for rawjump in jumpsToProcess:
        cleanedArgs = { k:v for (k,v) in rawjump.items() if v is not None }
        jump = StamperArgs(**cleanedArgs)
        cleaned_jumps.append(jump)
    return cleaned_jumps


def __main__():

    parser = jumpParser()
    cli_arg_list = getCmdLineList(parser)
    processJumps(cli_arg_list)  

    return 0


__main__()

